"""
Service de génération de rapports multi-formats.
"""
import io
import csv
import json
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from apps.products.models import Product, Category
from apps.stock_movements.models import Stock, StockMovement
from apps.purchase_orders.models import PurchaseOrder
from apps.inventories.models import Inventory
from apps.suppliers.models import Supplier
from apps.departments.models import Department
from apps.warehouses.models import Warehouse


class ReportService:
    """Service de génération de rapports."""
    
    # Couleurs IUC
    PRIMARY_COLOR = '#1e40af'
    SECONDARY_COLOR = '#166534'
    ACCENT_COLOR = '#ca8a04'
    LIGHT_GRAY = '#f3f4f6'
    WHITE = '#ffffff'
    DARK = '#111827'
    
    @staticmethod
    def generate_stock_level_report(format: str = 'PDF', filters: Dict = None) -> bytes:
        """Génère un rapport de niveau de stock."""
        queryset = Stock.objects.filter(is_deleted=False).select_related(
            'product__category', 'warehouse'
        )
        
        if filters:
            if filters.get('warehouse'):
                queryset = queryset.filter(warehouse_id=filters['warehouse'])
            if filters.get('category'):
                queryset = queryset.filter(product__category_id=filters['category'])
        
        data = []
        for stock in queryset:
            data.append({
                'product': stock.product.name,
                'reference': stock.product.reference,
                'category': stock.product.category.name if stock.product.category else '',
                'warehouse': stock.warehouse.name,
                'quantity': stock.quantity,
                'available': stock.available_quantity,
                'unit_price': float(stock.unit_price),
                'total_value': float(stock.quantity * stock.unit_price),
                'status': 'Rupture' if stock.quantity <= 0 else 'Bas' if stock.quantity <= stock.product.min_stock else 'Normal',
                'location': stock.location or '',
            })
        
        if format == 'PDF':
            return ReportService._generate_pdf(data, 'Niveau de Stock', 'Rapport de niveau de stock')
        elif format == 'EXCEL':
            return ReportService._generate_excel(data, 'Niveau de Stock')
        elif format == 'CSV':
            return ReportService._generate_csv(data)
        else:
            raise ValueError(f"Format {format} non supporté")
    
    @staticmethod
    def generate_movement_report(format: str = 'PDF', filters: Dict = None) -> bytes:
        """Génère un rapport de mouvements."""
        queryset = StockMovement.objects.filter(is_deleted=False).select_related(
            'stock__product', 'stock__warehouse', 'performed_by'
        ).order_by('-created_at')
        
        if filters:
            if filters.get('date_from'):
                queryset = queryset.filter(created_at__gte=filters['date_from'])
            if filters.get('date_to'):
                queryset = queryset.filter(created_at__lte=filters['date_to'])
            if filters.get('type'):
                queryset = queryset.filter(movement_type=filters['type'])
        
        data = []
        for mov in queryset[:1000]:
            data.append({
                'date': mov.created_at.strftime('%d/%m/%Y %H:%M'),
                'type': mov.get_movement_type_display(),
                'product': mov.stock.product.name,
                'warehouse': mov.stock.warehouse.name,
                'quantity': mov.quantity,
                'unit_price': float(mov.unit_price),
                'total': float(mov.total_price),
                'reason': mov.get_reason_display(),
                'user': mov.performed_by.get_full_name() if mov.performed_by else '',
                'document': mov.reference_document or '',
            })
        
        if format == 'PDF':
            return ReportService._generate_pdf(data, 'Mouvements de Stock', 'Rapport des mouvements')
        elif format == 'EXCEL':
            return ReportService._generate_excel(data, 'Mouvements')
        elif format == 'CSV':
            return ReportService._generate_csv(data)
        else:
            raise ValueError(f"Format {format} non supporté")
    
    @staticmethod
    def generate_inventory_report(inventory_id: str, format: str = 'PDF') -> bytes:
        """Génère un rapport d'inventaire."""
        try:
            inventory = Inventory.objects.get(id=inventory_id)
        except Inventory.DoesNotExist:
            raise ValueError("Inventaire introuvable")
        
        lines = inventory.lines.select_related('product').all()
        
        data = []
        for line in lines:
            data.append({
                'product': line.product.name,
                'reference': line.product.reference,
                'expected': line.expected_quantity,
                'counted': line.counted_quantity or 0,
                'difference': line.difference,
                'unit_price': float(line.unit_price),
                'diff_value': float(line.difference_value),
            })
        
        title = f"Inventaire {inventory.reference}"
        subtitle = f"{inventory.warehouse.name} - {inventory.start_date.strftime('%d/%m/%Y')}"
        
        if format == 'PDF':
            return ReportService._generate_pdf(data, title, subtitle)
        elif format == 'EXCEL':
            return ReportService._generate_excel(data, title)
        elif format == 'CSV':
            return ReportService._generate_csv(data)
        else:
            raise ValueError(f"Format {format} non supporté")
    
    @staticmethod
    def generate_consumption_report(format: str = 'PDF', filters: Dict = None) -> bytes:
        """Génère un rapport de consommation par département."""
        movements = StockMovement.objects.filter(
            is_deleted=False,
            movement_type='OUTPUT'
        ).select_related('stock__product', 'performed_by__department')
        
        if filters:
            if filters.get('date_from'):
                movements = movements.filter(created_at__gte=filters['date_from'])
            if filters.get('date_to'):
                movements = movements.filter(created_at__lte=filters['date_to'])
        
        # Regrouper par département
        dept_consumption = {}
        for mov in movements:
            dept = mov.performed_by.department if mov.performed_by and mov.performed_by.department else None
            dept_name = dept.name if dept else 'Non assigné'
            
            if dept_name not in dept_consumption:
                dept_consumption[dept_name] = {
                    'total_value': 0,
                    'total_items': 0,
                    'products': {}
                }
            
            dept_consumption[dept_name]['total_value'] += float(mov.total_price)
            dept_consumption[dept_name]['total_items'] += mov.quantity
            
            product_name = mov.stock.product.name
            if product_name not in dept_consumption[dept_name]['products']:
                dept_consumption[dept_name]['products'][product_name] = 0
            dept_consumption[dept_name]['products'][product_name] += mov.quantity
        
        data = []
        for dept_name, stats in dept_consumption.items():
            data.append({
                'department': dept_name,
                'total_value': stats['total_value'],
                'total_items': stats['total_items'],
                'top_products': ', '.join(
                    sorted(stats['products'].items(), key=lambda x: x[1], reverse=True)[:3]
                ),
            })
        
        if format == 'PDF':
            return ReportService._generate_pdf(data, 'Consommation par Département', 'Rapport de consommation')
        elif format == 'EXCEL':
            return ReportService._generate_excel(data, 'Consommation')
        elif format == 'CSV':
            return ReportService._generate_csv(data)
        else:
            raise ValueError(f"Format {format} non supporté")
    
    # ==========================================
    # GÉNÉRATEURS DE FORMATS
    # ==========================================
    
    @staticmethod
    def _generate_pdf(data: List[Dict], title: str, subtitle: str = '') -> bytes:
        """Génère un fichier PDF."""
        buffer = io.BytesIO()
        
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=15*mm,
            leftMargin=15*mm,
            topMargin=15*mm,
            bottomMargin=15*mm,
        )
        
        styles = getSampleStyleSheet()
        elements = []
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=16,
            textColor=ReportService.PRIMARY_COLOR,
            spaceAfter=2*mm,
        )
        elements.append(Paragraph(title, title_style))
        
        if subtitle:
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=10,
                textColor='#6b7280',
                spaceAfter=10*mm,
            )
            elements.append(Paragraph(subtitle, subtitle_style))
        
        # Date
        date_text = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        elements.append(Paragraph(date_text, styles['Normal']))
        elements.append(Spacer(1, 5*mm))
        
        if data:
            # En-têtes
            headers = list(data[0].keys())
            table_data = [headers]
            
            # Données
            for row in data:
                table_data.append([str(row.get(h, '')) for h in headers])
            
            # Créer le tableau
            col_widths = [doc.width / len(headers)] * len(headers)
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # Style du tableau
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), ReportService.PRIMARY_COLOR),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ])
            table.setStyle(style)
            elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def _generate_excel(data: List[Dict], sheet_name: str) -> bytes:
        """Génère un fichier Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        
        if not data:
            ws['A1'] = 'Aucune donnée'
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        
        # Styles
        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        header_fill = PatternFill(start_color=ReportService.PRIMARY_COLOR.replace('#', ''), end_color=ReportService.PRIMARY_COLOR.replace('#', ''), fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=9)
        cell_alignment = Alignment(vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )
        
        # En-têtes
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header.upper())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Données
        for row_num, row in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                value = row.get(header, '')
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                # Alternance couleurs
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
        
        # Ajuster largeur colonnes
        for col_num in range(1, len(headers) + 1):
            max_length = max(
                len(str(ws.cell(row=row_num, column=col_num).value or ''))
                for row_num in range(1, len(data) + 2)
            )
            ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 4, 40)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def _generate_csv(data: List[Dict]) -> bytes:
        """Génère un fichier CSV."""
        buffer = io.StringIO()
        
        if data:
            writer = csv.DictWriter(buffer, fieldnames=data[0].keys(), delimiter=';')
            writer.writeheader()
            writer.writerows(data)
        
        buffer.seek(0)
        return buffer.getvalue().encode('utf-8-sig')